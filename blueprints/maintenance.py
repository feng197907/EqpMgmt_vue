# 维护计划 Blueprint
from datetime import datetime
from io import BytesIO

from flask import Blueprint, flash, jsonify, redirect, render_template, request, send_file, url_for
from flask_login import current_user, login_required
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border

from config import FIXED_INTERVAL_LABELS, FIXED_INTERVAL_OPTIONS, MAINTENANCE_RESULTS, MAINTENANCE_RESULT_LABELS, MAINTENANCE_TYPE_LABELS, MAINTENANCE_TYPES
from database import get_db
from models.electronic_signature import ElectronicSignature, SIGN_MEANINGS
from models.maintenance import DeviceRepairRecord, MaintenancePlan, MaintenanceRecord
from utils.audit import log_action
from utils.decorators import admin_required, permission_required
from utils.maintenance import calc_next_due_date, calc_urgency

maintenance_bp = Blueprint("maintenance", __name__, url_prefix="/device/<int:device_id>/maintenance")


@maintenance_bp.route("/")
@login_required
def maintenance_plans(device_id):
    """设备维护计划列表页"""
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM devices WHERE id = %s", (device_id,))
    device = cur.fetchone()
    if device is None:
        conn.close()
        flash("设备不存在。", "warning")
        return redirect(url_for("auth.index"))

    plans = MaintenancePlan.get_by_device(device_id, active_only=True)
    conn.close()

    # 为每个计划计算紧迫度
    plan_list = []
    for plan in plans:
        plan_dict = {
            "plan": plan,
            "urgency": calc_urgency(plan.next_due_date),
            "overdue_days": plan.overdue_days if plan.overdue_days < 0 else 0,
        }
        plan_list.append(plan_dict)

    return render_template(
        "device_maintenance.html",
        device=device,
        plans=plan_list,
        maintenance_types=MAINTENANCE_TYPES,
        type_labels=MAINTENANCE_TYPE_LABELS,
        fixed_intervals=FIXED_INTERVAL_OPTIONS,
        interval_labels=FIXED_INTERVAL_LABELS,
    )


@maintenance_bp.route("/plan", methods=["POST"])
@permission_required("device_maintenance")
def create_plan(device_id):
    """添加维护计划"""
    maintenance_type = request.form.get("maintenance_type", "").strip()
    interval_days = request.form.get("interval_days", "").strip()
    first_due_date = request.form.get("first_due_date", "").strip()

    if not maintenance_type or not interval_days or not first_due_date:
        flash("请填写完整的维护计划信息。", "warning")
        return redirect(url_for("maintenance.maintenance_plans", device_id=device_id))

    try:
        interval_days = int(interval_days)
        if interval_days < 1 or interval_days > 365:
            flash("周期天数必须在1-365之间。", "warning")
            return redirect(url_for("maintenance.maintenance_plans", device_id=device_id))
    except ValueError:
        flash("周期天数必须是有效数字。", "warning")
        return redirect(url_for("maintenance.maintenance_plans", device_id=device_id))

    # 检查是否已存在相同类型的激活计划
    existing = MaintenancePlan.get_by_device_and_type(device_id, maintenance_type)
    if existing:
        flash("该设备已存在相同类型的激活维护计划。", "warning")
        return redirect(url_for("maintenance.maintenance_plans", device_id=device_id))

    # 创建新计划
    plan = MaintenancePlan(
        device_id=device_id,
        maintenance_type=maintenance_type,
        interval_days=interval_days,
        next_due_date=first_due_date,
        is_active=1,
        created_by=current_user.username,
    )
    plan.save()

    log_action(
        current_user.username, "create_maintenance_plan", "maintenance_plan", plan.id,
        f"创建设备 {device_id} 的{MAINTENANCE_TYPE_LABELS.get(maintenance_type, '')}计划",
    )

    flash("维护计划已创建。", "success")
    return redirect(url_for("maintenance.maintenance_plans", device_id=device_id))


@maintenance_bp.route("/plan/<int:plan_id>", methods=["PUT"])
@permission_required("device_maintenance")
def update_plan(device_id, plan_id):
    """编辑维护计划（API接口）"""
    data = request.get_json()
    if not data:
        return jsonify({"error": "无效的请求数据"}), 400

    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM maintenance_plan WHERE id = %s AND device_id = %s", (plan_id, device_id))
    row = cur.fetchone()
    if row is None:
        conn.close()
        return jsonify({"error": "维护计划不存在"}), 404

    old_values = dict(row)
    plan = MaintenancePlan(**old_values)

    # 更新字段
    if "interval_days" in data:
        interval_days = int(data["interval_days"])
        if interval_days < 1 or interval_days > 365:
            return jsonify({"error": "周期天数必须在1-365之间"}), 400
        plan.interval_days = interval_days

    if "next_due_date" in data:
        plan.next_due_date = data["next_due_date"]

    if "is_active" in data:
        plan.is_active = 1 if data["is_active"] else 0

    plan.save()
    conn.close()

    log_action(
        current_user.username, "update_maintenance_plan", "maintenance_plan", plan_id,
        "更新维护计划",
        before_value={"interval_days": old_values["interval_days"], "next_due_date": old_values["next_due_date"]},
        after_value={"interval_days": plan.interval_days, "next_due_date": plan.next_due_date},
    )

    return jsonify({
        "id": plan.id,
        "message": "维护计划已更新",
    })


@maintenance_bp.route("/plan/<int:plan_id>", methods=["DELETE"])
@permission_required("device_maintenance")
def delete_plan(device_id, plan_id):
    """删除维护计划（软删除）"""
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM maintenance_plan WHERE id = %s AND device_id = %s", (plan_id, device_id))
    row = cur.fetchone()
    if row is None:
        conn.close()
        return jsonify({"error": "维护计划不存在"}), 404

    # 软删除：设置is_active=0
    cur.execute("UPDATE maintenance_plan SET is_active = 0, updated_at = NOW() WHERE id = %s", (plan_id,))
    conn.commit()
    conn.close()

    log_action(
        current_user.username, "delete_maintenance_plan", "maintenance_plan", plan_id,
        "删除维护计划（软删除）",
    )

    return jsonify({"message": "维护计划已删除"})


@maintenance_bp.route("/plan/<int:plan_id>/close", methods=["POST"])
@permission_required("device_maintenance")
def close_plan(device_id, plan_id):
    """关闭维护计划（仅逾期计划可关闭）"""
    data = request.get_json(silent=True) or {}
    close_reason = data.get("close_reason", "").strip()

    plan = MaintenancePlan.get_by_id(plan_id)
    if plan is None or plan.device_id != device_id:
        return jsonify({"error": "维护计划不存在"}), 404

    if not plan.is_active:
        return jsonify({"error": "维护计划已删除，无法关闭"}), 400

    if plan.is_closed:
        return jsonify({"error": "维护计划已关闭"}), 400

    if not plan.is_overdue:
        return jsonify({"error": "仅逾期的维护计划可以关闭"}), 400

    if not close_reason:
        return jsonify({"error": "请填写关闭原因"}), 400

    plan.close(current_user.username, close_reason)

    log_action(
        current_user.username, "close_maintenance_plan", "maintenance_plan", plan_id,
        f"关闭逾期维护计划：{plan.maintenance_type_label}，原因：{close_reason}",
    )

    return jsonify({"message": "维护计划已关闭"})


@maintenance_bp.route("/plan/<int:plan_id>/record", methods=["GET"])
@login_required
def new_record_form(device_id, plan_id):
    """获取维护记录表单"""
    conn = get_db()
    cur = conn.cursor()

    # 获取维护计划信息
    cur.execute("SELECT * FROM maintenance_plan WHERE id = %s AND device_id = %s AND is_active = 1", (plan_id, device_id))
    plan = cur.fetchone()
    if plan is None:
        conn.close()
        flash("维护计划不存在或已停用。", "warning")
        return redirect(url_for("maintenance.maintenance_plans", device_id=device_id))

    # 获取设备信息
    cur.execute("SELECT * FROM devices WHERE id = %s", (device_id,))
    device = cur.fetchone()
    if device is None:
        conn.close()
        flash("设备不存在。", "warning")
        return redirect(url_for("auth.index"))

    conn.close()

    return render_template(
        "maintenance_record_form.html",
        device=device,
        plan=plan,
        maintenance_type_label=MAINTENANCE_TYPE_LABELS.get(plan["maintenance_type"], ""),
        results=MAINTENANCE_RESULTS,
        result_labels=MAINTENANCE_RESULT_LABELS,
    )


@maintenance_bp.route("/plan/<int:plan_id>/record", methods=["POST"])
@permission_required("device_maintenance")
def submit_record(device_id, plan_id):
    """提交维护记录"""
    content = request.form.get("content", "").strip()
    result = request.form.get("result", "").strip()
    parts_used = request.form.get("parts_used", "").strip()

    if not content or not result:
        flash("请填写维护内容和结果。", "warning")
        return redirect(url_for("maintenance.new_record_form", device_id=device_id, plan_id=plan_id))

    conn = get_db()
    cur = conn.cursor()

    # 获取维护计划
    cur.execute("SELECT * FROM maintenance_plan WHERE id = %s AND device_id = %s AND is_active = 1", (plan_id, device_id))
    plan = cur.fetchone()
    if plan is None:
        conn.close()
        flash("维护计划不存在或已停用。", "warning")
        return redirect(url_for("maintenance.maintenance_plans", device_id=device_id))

    # 获取设备信息
    cur.execute("SELECT * FROM devices WHERE id = %s", (device_id,))
    device = cur.fetchone()
    if device is None:
        conn.close()
        flash("设备不存在。", "warning")
        return redirect(url_for("auth.index"))

    # 计算本次维护时间（使用当前时间）
    performed_at = datetime.now().strftime("%Y-%m-%d")
    next_due_date = calc_next_due_date(performed_at, plan["interval_days"])

    # 创建维护记录
    record = MaintenanceRecord(
        plan_id=plan_id,
        device_id=device_id,
        maintenance_type=plan["maintenance_type"],
        content=content,
        result=result,
        performed_by=current_user.username,
        performed_at=performed_at,
        next_due_date=next_due_date,
        parts_used=parts_used if parts_used else None,
    )
    record.save()

    # 为执行人自动生成电子签名（21 CFR Part 11 执行确认）
    ip_address = request.headers.get("X-Forwarded-For", request.remote_addr)
    esign = ElectronicSignature(
        record_type="maintenance_record",
        record_id=record.id,
        signed_by=current_user.username,
        signed_by_display=current_user.role_label or current_user.username,
        sign_meaning="executed",
        sign_meaning_label=SIGN_MEANINGS["executed"],
        ip_address=ip_address,
        remark="维护记录提交：{}".format(MAINTENANCE_RESULT_LABELS.get(result, result)),
    )
    esign.save()

    # 根据维护结果决定是否更新到期日
    if result == "qualified":
        cur.execute(
            "UPDATE maintenance_plan SET next_due_date = %s, updated_at = NOW() WHERE id = %s",
            (next_due_date, plan_id)
        )
        conn.commit()
        message = "维护记录已保存，下次到期日已更新"
    else:
        message = "维护记录已保存（结果为不合格/待处理，到期日未更新）"
        conn.commit()

    log_action(
        current_user.username, "submit_maintenance_record", "maintenance_record", record.id,
        f"提交设备 {device['device_code']} 的维护记录，结果：{result}",
    )

    conn.close()

    flash(message, "success")
    return redirect(url_for("maintenance.maintenance_plans", device_id=device_id))


@maintenance_bp.route("/record/<int:record_id>/delete", methods=["POST"])
@admin_required
def delete_record(device_id, record_id):
    """删除维护记录（管理员）"""
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        """SELECT mr.id, mr.device_id, mr.plan_id, mr.maintenance_type, mr.content, mr.result,
                  mp.maintenance_type AS plan_type
           FROM maintenance_record mr
           JOIN maintenance_plan mp ON mp.id = mr.plan_id
           WHERE mr.id = %s AND mr.device_id = %s""",
        (record_id, device_id),
    )
    record = cur.fetchone()
    if record is None:
        conn.close()
        flash("维护记录不存在。", "warning")
        return redirect(url_for("maintenance.maintenance_history", device_id=device_id))

    cur.execute("DELETE FROM maintenance_record WHERE id = %s", (record_id,))
    conn.commit()

    log_action(
        current_user.username,
        "delete_maintenance_record",
        "maintenance_record",
        record_id,
        f"删除设备 {device_id} 的维护记录：{record['content']}（结果：{record['result']}）",
    )
    conn.close()
    flash("维护记录已删除。", "success")
    return redirect(url_for("maintenance.maintenance_history", device_id=device_id))


@maintenance_bp.route("/history", methods=["GET"])
@login_required
def maintenance_history(device_id):
    """设备维护历史列表页"""
    conn = get_db()
    cur = conn.cursor()

    # 获取设备信息
    cur.execute("SELECT * FROM devices WHERE id = %s", (device_id,))
    device = cur.fetchone()
    if device is None:
        conn.close()
        flash("设备不存在。", "warning")
        return redirect(url_for("auth.index"))

    # 获取筛选参数
    maintenance_type = request.args.get("type", "").strip()
    year = request.args.get("year", "").strip()
    try:
        page = max(1, int(request.args.get("page", 1)))
    except ValueError:
        page = 1
    per_page = 20

    # 获取维护记录（分页）
    records, pagination = MaintenanceRecord.get_by_device(
        device_id, maintenance_type=maintenance_type if maintenance_type else None,
        year=year if year else None, page=page, per_page=per_page
    )

    display_records = []
    for record in records:
        if record.result == "qualified":
            result_text = "合格"
            result_class = "bg-success"
        elif record.result == "unqualified":
            result_text = "不合格"
            result_class = "bg-danger"
        elif record.result == "pending":
            result_text = "待处理"
            result_class = "bg-warning text-dark"
        else:
            result_text = record.result if record.result else "-"
            result_class = "bg-secondary"
        display_records.append(
            {
                "id": record.id,
                "performed_at": record.performed_at,
                "maintenance_type_label": record.maintenance_type_label,
                "content": record.content,
                "result_text": result_text,
                "result_class": result_class,
                "performed_by": record.performed_by,
                "next_due_date": record.next_due_date,
            }
        )

    conn.close()

    return render_template(
        "device_maintenance_history.html",
        device=device,
        records=display_records,
        pagination=pagination,
        maintenance_types=MAINTENANCE_TYPES,
        type_labels=MAINTENANCE_TYPE_LABELS,
        results=MAINTENANCE_RESULTS,
        result_labels=MAINTENANCE_RESULT_LABELS,
        selected_type=maintenance_type,
        selected_year=year,
    )


# ============================================================
# API 接口
# ============================================================

@maintenance_bp.route("/api/plans", methods=["GET"])
@login_required
def api_get_plans(device_id):
    """获取设备维护计划列表（API）"""
    plans = MaintenancePlan.get_by_device(device_id, active_only=True)
    result = []
    for plan in plans:
        result.append({
            "id": plan.id,
            "device_id": plan.device_id,
            "maintenance_type": plan.maintenance_type,
            "maintenance_type_label": plan.maintenance_type_label,
            "interval_days": plan.interval_days,
            "next_due_date": plan.next_due_date,
            "is_active": bool(plan.is_active),
            "created_by": plan.created_by,
            "created_at": plan.created_at,
            "overdue_days": plan.overdue_days if plan.overdue_days < 0 else 0,
            "urgency": calc_urgency(plan.next_due_date),
        })
    return jsonify({"plans": result})


@maintenance_bp.route("/api/records", methods=["GET"])
@login_required
def api_get_records(device_id):
    """获取设备维护记录列表（API）"""
    maintenance_type = request.args.get("type", "").strip()
    year = request.args.get("year", "").strip()
    try:
        page = max(1, int(request.args.get("page", 1)))
    except ValueError:
        page = 1
    per_page = int(request.args.get("per_page", 20))

    records, pagination = MaintenanceRecord.get_by_device(
        device_id, maintenance_type=maintenance_type if maintenance_type else None,
        year=year if year else None, page=page, per_page=per_page
    )

    result = []
    for record in records:
        result.append({
            "id": record.id,
            "plan_id": record.plan_id,
            "device_id": record.device_id,
            "maintenance_type": record.maintenance_type,
            "maintenance_type_label": record.maintenance_type_label,
            "content": record.content,
            "result": record.result,
            "performed_by": record.performed_by,
            "performed_at": record.performed_at,
            "next_due_date": record.next_due_date,
            "parts_used": record.parts_used,
            "created_at": record.created_at,
        })

    return jsonify({"records": result, "pagination": pagination})


# ============================================================
# 维修记录（独立于维护计划）
# ============================================================

@maintenance_bp.route("/repair", methods=["GET"])
@login_required
def repair_records(device_id):
    """设备维修记录列表页"""
    conn = get_db()
    cur = conn.cursor()

    # 获取设备信息
    cur.execute("SELECT * FROM devices WHERE id = %s", (device_id,))
    device = cur.fetchone()
    if device is None:
        conn.close()
        flash("设备不存在。", "warning")
        return redirect(url_for("auth.index"))

    try:
        page = max(1, int(request.args.get("page", 1)))
    except ValueError:
        page = 1
    per_page = 20

    # 获取维修记录（分页）
    records, pagination = DeviceRepairRecord.get_by_device(device_id, page=page, per_page=per_page)

    display_records = []
    for record in records:
        if record.result == "qualified":
            result_text = "合格"
            result_class = "bg-success"
        elif record.result == "unqualified":
            result_text = "不合格"
            result_class = "bg-danger"
        elif record.result == "pending":
            result_text = "待处理"
            result_class = "bg-warning text-dark"
        else:
            result_text = record.result if record.result else "-"
            result_class = "bg-secondary"
        display_records.append(
            {
                "id": record.id,
                "performed_at": record.performed_at,
                "content": record.content,
                "result_text": result_text,
                "result_class": result_class,
                "performed_by": record.performed_by,
                "parts_used": record.parts_used,
            }
        )

    conn.close()

    return render_template(
        "device_repair_records.html",
        device=device,
        records=display_records,
        pagination=pagination,
        results=MAINTENANCE_RESULTS,
        result_labels=MAINTENANCE_RESULT_LABELS,
    )


@maintenance_bp.route("/repair/new", methods=["GET"])
@login_required
def new_repair_form(device_id):
    """新增维修记录表单页"""
    conn = get_db()
    cur = conn.cursor()

    # 获取设备信息
    cur.execute("SELECT * FROM devices WHERE id = %s", (device_id,))
    device = cur.fetchone()
    if device is None:
        conn.close()
        flash("设备不存在。", "warning")
        return redirect(url_for("auth.index"))

    conn.close()

    return render_template(
        "repair_record_form.html",
        device=device,
        results=MAINTENANCE_RESULTS,
        result_labels=MAINTENANCE_RESULT_LABELS,
    )


@maintenance_bp.route("/repair/new", methods=["POST"])
@permission_required("device_maintenance")
def submit_repair_record(device_id):
    """提交维修记录"""
    content = request.form.get("content", "").strip()
    result = request.form.get("result", "").strip()
    parts_used = request.form.get("parts_used", "").strip()
    performed_at = request.form.get("performed_at", "").strip()

    if not content or not result:
        flash("请填写维修内容和结果。", "warning")
        return redirect(url_for("maintenance.new_repair_form", device_id=device_id))

    conn = get_db()
    cur = conn.cursor()

    # 获取设备信息
    cur.execute("SELECT * FROM devices WHERE id = %s", (device_id,))
    device = cur.fetchone()
    if device is None:
        conn.close()
        flash("设备不存在。", "warning")
        return redirect(url_for("auth.index"))

    # 使用表单日期或当前时间
    if not performed_at:
        performed_at = datetime.now().strftime("%Y-%m-%d")

    # 创建维修记录
    record = DeviceRepairRecord(
        device_id=device_id,
        content=content,
        result=result,
        performed_by=current_user.username,
        performed_at=performed_at,
        parts_used=parts_used if parts_used else None,
    )
    record.save()

    log_action(
        current_user.username, "submit_repair_record", "repair_record", record.id,
        f"提交设备 {device['device_code']} 的维修记录，结果：{result}",
    )

    conn.close()

    flash("维修记录已保存。", "success")
    return redirect(url_for("maintenance.repair_records", device_id=device_id))


@maintenance_bp.route("/repair/<int:record_id>/delete", methods=["POST"])
@admin_required
def delete_repair_record(device_id, record_id):
    """删除维修记录（管理员）"""
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        "SELECT * FROM repair_record WHERE id = %s AND device_id = %s",
        (record_id, device_id),
    )
    record = cur.fetchone()
    if record is None:
        conn.close()
        flash("维修记录不存在。", "warning")
        return redirect(url_for("maintenance.repair_records", device_id=device_id))

    cur.execute("DELETE FROM repair_record WHERE id = %s", (record_id,))
    conn.commit()

    log_action(
        current_user.username,
        "delete_repair_record",
        "repair_record",
        record_id,
        f"删除设备 {device_id} 的维修记录：{record['content']}（结果：{record['result']}）",
    )
    conn.close()
    flash("维修记录已删除。", "success")
    return redirect(url_for("maintenance.repair_records", device_id=device_id))


# ============================================================
# 导出功能
# ============================================================

def _create_excel_response(wb, filename):
    """将 Workbook 转为 Flask send_file 响应"""
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


def _set_header_style(cell):
    """设置表头样式"""
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _set_border(cell):
    """设置单元格边框"""
    thin = Side(style="thin", color="000000")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


@maintenance_bp.route("/export/plans")
@login_required
def export_plans(device_id):
    """导出维护计划为 Excel"""
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT device_code, device_name FROM devices WHERE id = %s", (device_id,))
    device = cur.fetchone()
    conn.close()

    if device is None:
        flash("设备不存在。", "warning")
        return redirect(url_for("auth.index"))

    plans = MaintenancePlan.get_by_device(device_id, active_only=False)

    wb = Workbook()
    ws = wb.active
    ws.title = "维护计划"

    headers = ["维护类型", "周期天数", "下次到期日", "状态", "创建人", "创建时间", "更新时间"]
    ws.append(headers)
    for cell in ws[1]:
        _set_header_style(cell)
        _set_border(cell)

    for plan in plans:
        ws.append([
            plan.maintenance_type_label,
            plan.interval_days,
            plan.next_due_date,
            "激活" if plan.is_active else "停用",
            plan.created_by or "-",
            plan.created_at or "-",
            plan.updated_at or "-",
        ])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            _set_border(cell)
            cell.alignment = Alignment(vertical="center")

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    filename = f"{device['device_code']}_维护计划_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return _create_excel_response(wb, filename)


@maintenance_bp.route("/export/history")
@login_required
def export_history(device_id):
    """导出维护历史为 Excel"""
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT device_code, device_name FROM devices WHERE id = %s", (device_id,))
    device = cur.fetchone()
    conn.close()

    if device is None:
        flash("设备不存在。", "warning")
        return redirect(url_for("auth.index"))

    maintenance_type = request.args.get("type", "").strip()
    year = request.args.get("year", "").strip()

    records, _ = MaintenanceRecord.get_by_device(
        device_id,
        maintenance_type=maintenance_type if maintenance_type else None,
        year=year if year else None,
        page=1,
        per_page=9999,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "维护历史"

    headers = ["维护时间", "维护类型", "维护内容", "结果", "执行人", "下次到期日", "备件使用"]
    ws.append(headers)
    for cell in ws[1]:
        _set_header_style(cell)
        _set_border(cell)

    for record in records:
        result_text = MAINTENANCE_RESULT_LABELS.get(record.result, record.result)
        ws.append([
            record.performed_at,
            record.maintenance_type_label,
            record.content,
            result_text,
            record.performed_by,
            record.next_due_date or "-",
            record.parts_used or "-",
        ])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            _set_border(cell)
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    filename = f"{device['device_code']}_维护历史_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return _create_excel_response(wb, filename)


@maintenance_bp.route("/export/repair")
@login_required
def export_repair(device_id):
    """导出维修记录为 Excel"""
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT device_code, device_name FROM devices WHERE id = %s", (device_id,))
    device = cur.fetchone()
    conn.close()

    if device is None:
        flash("设备不存在。", "warning")
        return redirect(url_for("auth.index"))

    records, _ = DeviceRepairRecord.get_by_device(device_id, page=1, per_page=9999)

    wb = Workbook()
    ws = wb.active
    ws.title = "维修记录"

    headers = ["维修时间", "维修内容", "结果", "执行人", "备件使用", "创建时间"]
    ws.append(headers)
    for cell in ws[1]:
        _set_header_style(cell)
        _set_border(cell)

    for record in records:
        result_text = MAINTENANCE_RESULT_LABELS.get(record.result, record.result)
        ws.append([
            record.performed_at,
            record.content,
            result_text,
            record.performed_by,
            record.parts_used or "-",
            record.created_at or "-",
        ])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            _set_border(cell)
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    filename = f"{device['device_code']}_维修记录_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return _create_excel_response(wb, filename)
